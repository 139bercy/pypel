import abc
import pandas as pd
import re
import openpyxl
from pypel.utils.utils import arrayer
import warnings
import logging
from pypel.config.config import get_config
from typing import Dict, Optional, List, Union, Any

logger = logging.getLogger(__name__)
logger.setLevel(getattr(logging, get_config()["LOGS_LEVEL"]))


class BaseExtractor:
    @abc.abstractmethod
    def extract(self, *args, **kwargs) -> Any:
        """This method must be implemented"""


class Extractor(BaseExtractor):
    """
    Encapsulates all the extracting, getting data logic.
    """
    def extract(self, file_path: Union[str, bytes],
                converters: Optional[Dict[str, type]] = None,
                dates: Optional[List[str]] = False,
                sheet_name: Union[None, int, str, List[Union[int, str]]] = 0,
                skiprows: Optional[int] = None, **kwargs) -> pd.DataFrame:
        """
        Read the passed file and return it as a dataframe.
        Uses many pandas parameters defined at Extractor instanciation.

        :param file_path: absolute path to the file
        :param converters:
            cf [pandas' doc](https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_excel.html)
        :param dates: this is equivalent to pandas' `parse_dates` in
             [read_excel](https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_excel.html)
        :param sheet_name:
            cf [pandas' doc](https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_excel.html)
        :param skiprows:
            cf [pandas' doc](https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_excel.html)
        :param kwargs: additional pandas args
        :return: pandas.Dataframe object
        """
        if file_path.endswith(".csv"):
            if get_config()["LOGS"]:
                with open(file_path) as file:
                    row_count = sum(1 for row in file)
                file_name = re.findall(r"(?<=/)[^/]*$", file_path)[0]
                logger.debug(f"{row_count} rows (including header)"
                             f"detected in the csv {file_name}")
            return pd.read_csv(file_path,
                               converters=converters,
                               parse_dates=dates,
                               **kwargs)
        elif file_path.endswith(".xlsx"):
            if get_config()["LOGS"]:
                wb = openpyxl.load_workbook(filename=file_path)
                if sheet_name != 0:
                    sheet = wb[sheet_name]
                else:
                    sheet = wb[wb.sheetnames[0]]
                    sheet_name = sheet.title
                excel_rows = sheet.max_row
                try:
                    file_name = re.findall(r"(?<=/)[.\s\w_-]+$", file_path)[0]
                except IndexError:
                    warnings.warn(f"Could not get file name from file path :"
                                  f"{file_path}")
                    file_name = "ERROR"
                logger.debug(f"{excel_rows} rows in the excel sheet \'{sheet_name}\'   from file \'{file_name}\'")
            if skiprows is not None:
                skiprows = arrayer(skiprows)
            return pd.read_excel(io=file_path,
                                 skiprows=skiprows,
                                 sheet_name=sheet_name,
                                 converters=converters,
                                 **kwargs)
        elif file_path.endswith(".xls"):
            if get_config()["LOGS"]:
                logger.debug(f"Targeting file \'{file_path}\'")
            if skiprows is not None:
                skiprows = arrayer(skiprows)
            return pd.read_excel(io=file_path,
                                 skiprows=skiprows,
                                 sheet_name=sheet_name,
                                 converters=converters,
                                 engine="xlrd",
                                 **kwargs)
        else:
            raise ValueError("File has unsupported file extension")


class CSVExtractor(BaseExtractor):
    def extract(self, file_path: Union[str, bytes],
                converters: Optional[Dict[str, type]] = None,
                dates: Optional[List[str]] = False,
                skiprows: Optional[int] = None, **kwargs) -> pd.DataFrame:
        """
        Read the passed file and return it as a dataframe.
        Uses many pandas parameters defined at Extractor instanciation.

        :param file_path: absolute path to the file
        :param converters:
            cf [pandas' doc](https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_excel.html)
        :param dates: this is equivalent to pandas' `parse_dates` in
             [read_excel](https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_excel.html)
        :param skiprows:
            cf [pandas' doc](https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_excel.html)
        :param kwargs: additional pandas args
        :return: pandas.Dataframe object
        """
        if get_config()["LOGS"]:
            with open(file_path) as file:
                row_count = sum(1 for row in file)
            file_name = re.findall(r"(?<=/)[^/]*$", file_path)[0]
            logger.debug(f"{row_count} rows (including header)"
                         f"detected in the csv {file_name}")
        return pd.read_csv(file_path,
                           converters=converters,
                           parse_dates=dates,
                           **kwargs)


class XLSExtractor(BaseExtractor):
    def extract(self, file_path: Union[str, bytes],
                converters: Optional[Dict[str, type]] = None,
                dates: Optional[List[str]] = False,
                sheet_name: Union[None, int, str, List[Union[int, str]]] = 0,
                skiprows: Optional[int] = None, **kwargs) -> pd.DataFrame:
        """
        Read the passed file and return it as a dataframe.
        Uses many pandas parameters defined at Extractor instanciation.

        :param file_path: absolute path to the file
        :param converters:
            cf [pandas' doc](https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_excel.html)
        :param dates: this is equivalent to pandas' `parse_dates` in
             [read_excel](https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_excel.html)
        :param sheet_name:
            cf [pandas' doc](https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_excel.html)
        :param skiprows:
            cf [pandas' doc](https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_excel.html)
        :param kwargs: additional pandas args
        :return: pandas.Dataframe object
        """
        if get_config()["LOGS"]:
            logger.debug(f"Targeting file \'{file_path}\'")
        if skiprows is not None:
            skiprows = arrayer(skiprows)
        return pd.read_excel(io=file_path,
                             skiprows=skiprows,
                             sheet_name=sheet_name,
                             converters=converters,
                             engine="xlrd",
                             **kwargs)


class XLSXExtractor(BaseExtractor):
    def extract(self, file_path: Union[str, bytes],
                converters: Optional[Dict[str, type]] = None,
                dates: Optional[List[str]] = False,
                sheet_name: Union[None, int, str, List[Union[int, str]]] = 0,
                skiprows: Optional[int] = None, **kwargs) -> pd.DataFrame:
        """
        Read the passed file and return it as a dataframe.
        Uses many pandas parameters defined at Extractor instanciation.

        :param file_path: absolute path to the file
        :param converters:
            cf [pandas' doc](https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_excel.html)
        :param dates: this is equivalent to pandas' `parse_dates` in
             [read_excel](https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_excel.html)
        :param sheet_name:
            cf [pandas' doc](https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_excel.html)
        :param skiprows:
            cf [pandas' doc](https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_excel.html)
        :param kwargs: additional pandas args
        :return: pandas.Dataframe object
        """
        if get_config()["LOGS"]:
            wb = openpyxl.load_workbook(filename=file_path)
            if sheet_name != 0:
                sheet = wb[sheet_name]
            else:
                sheet = wb[wb.sheetnames[0]]
                sheet_name = sheet.title
            excel_rows = sheet.max_row
            try:
                file_name = re.findall(r"(?<=/)[.\s\w_-]+$", file_path)[0]
            except IndexError:
                warnings.warn(f"Could not get file name from file path :"
                              f"{file_path}")
                file_name = "ERROR"
            logger.debug(f"{excel_rows} rows in the excel sheet \'{sheet_name}\'   from file \'{file_name}\'")
        if skiprows is not None:
            skiprows = arrayer(skiprows)
        return pd.read_excel(io=file_path,
                             skiprows=skiprows,
                             sheet_name=sheet_name,
                             converters=converters,
                             **kwargs)
