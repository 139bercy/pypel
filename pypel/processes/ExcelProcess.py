# -*- coding: UTF-8 -*-
import pandas as pd
import openpyxl
from pypel.processes.Process import Process
import logging.handlers
import re


logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)


class ExcelProcess(Process):

    def init_dataframe(self, file_path: str, sheet_name: str = 0, skiprows=None):
        """
        Read the passed file and return it as a dataframe.
        Uses pandas.read_excel's `converters`, `parse_dates`, `sheet_name` and `skiprows` parameters.

        :param file_path: absolute path to the file
        :param sheet_name: the excel sheet name to load
        :param skiprows: the number of lines to skip at the top of the file
        :return: a pandas.Dataframe object
        """
        if skiprows is None:
            skiprows = skiprows
        else:
            skiprows = ExcelProcess.arrayer(skiprows)
        if file_path.endswith(".csv"):
            return super().init_dataframe(file_path)
        else:
            wb = openpyxl.load_workbook(filename=file_path)
            if sheet_name != 0:
                sheet = wb[sheet_name]
            else:
                sheet = wb[wb.sheetnames[0]]
            excel_rows = sheet.max_row
            try:
                file_name = re.findall(r"(?<=/)[.\s\w_-]+$", file_path)[0]
            except IndexError:
                logger.error(f"Could not get file name from file path : {file_path}")
                file_name = "ERROR"
            logger.info(f"{excel_rows} rows in the excel sheet \'{sheet_name}\' from file \'{file_name}\'")
            return pd.read_excel(io=file_path,
                                 skiprows=skiprows,
                                 sheet_name=sheet_name,
                                 converters=self.converters,
                                 parse_dates=self.dates)

    @staticmethod
    def arrayer(integer: int):
        """Return a list of integers from 0 to `integer`, inclusive"""
        return [x for x in range(integer + 1)]
